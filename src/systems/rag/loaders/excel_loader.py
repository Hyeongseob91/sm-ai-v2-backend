"""Excel 파일 로더 - openpyxl 기반"""

from typing import List
from langchain_core.documents import Document
from openpyxl import load_workbook


class ExcelLoader:
    """
    openpyxl을 사용한 Excel 파일 로더.
    각 시트를 별도의 Document로 변환하며, 마크다운 테이블 형식으로 출력합니다.
    """

    def __init__(self, file_path: str):
        self.file_path = file_path

    def load(self) -> List[Document]:
        """Excel 파일을 로드하여 Document 리스트로 반환합니다."""
        documents = []

        workbook = load_workbook(
            filename=self.file_path, read_only=True, data_only=True
        )

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            content = self._sheet_to_markdown_table(sheet)

            if content.strip():
                doc = Document(
                    page_content=content,
                    metadata={
                        "source": self.file_path,
                        "sheet_name": sheet_name,
                        "file_type": "xlsx",
                    },
                )
                documents.append(doc)

        workbook.close()
        return documents

    def _sheet_to_markdown_table(self, sheet) -> str:
        """시트 데이터를 마크다운 테이블로 변환합니다."""
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            return ""

        # 빈 행 제거
        non_empty_rows = [
            row for row in rows if any(cell is not None for cell in row)
        ]

        if not non_empty_rows:
            return ""

        # 첫 행을 헤더로 처리
        header = non_empty_rows[0]
        header_values = [str(cell) if cell is not None else "" for cell in header]

        lines = []
        lines.append("| " + " | ".join(header_values) + " |")
        lines.append("| " + " | ".join(["---"] * len(header)) + " |")

        for row in non_empty_rows[1:]:
            row_values = [str(cell) if cell is not None else "" for cell in row]
            lines.append("| " + " | ".join(row_values) + " |")

        return "\n".join(lines)
